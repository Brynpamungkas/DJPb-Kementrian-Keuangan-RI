import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

# Pengaturan File (letakkan script di dalam folder yang sama dengan file Excel)
namaFileMonitoring = "Monitoring Retur SP2D_1783646455899_107.xlsx"
sheetMonitoring = "Monitoring Retur SP2D - SPANEXT"
namaFileSurat = "Surat Retur S-716.xlsx"
namaFileOutput = "Hasil_Automatis_Baru.xlsx"

# Deteksi lokasi folder otomatis
baseDir = os.path.dirname(os.path.abspath(__file__)) or os.getcwd()
fileMonitoring = os.path.join(baseDir, namaFileMonitoring)
fileSurat = os.path.join(baseDir, namaFileSurat)
fileOutput = os.path.join(baseDir, namaFileOutput)

def main():
    print(f"Membaca data dari:\n1. {fileMonitoring}\n2. {fileSurat}\n...")
    
    df1 = pd.read_excel(fileMonitoring, sheet_name=sheetMonitoring)
    df1 = df1.dropna(subset=['NO SP2D'])
    
    xl2 = pd.ExcelFile(fileSurat)
    df2List = []
    for sheet in xl2.sheet_names:
        tempDf = pd.read_excel(fileSurat, sheet_name=sheet)
        if 'NomorSP2D' in tempDf.columns:
            tempDf = tempDf.rename(columns={'NomorSP2D': 'Nomor SP2D'})
        df2List.append(tempDf)
    
    df2 = pd.concat(df2List, ignore_index=True)
    df2 = df2.dropna(subset=['Nomor SP2D'])
    
    df1['NO SP2D_clean'] = df1['NO SP2D'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    df2['Nomor SP2D_clean'] = df2['Nomor SP2D'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    
    sp2d1 = df1['NO SP2D_clean'].unique()
    sp2d2 = df2['Nomor SP2D_clean'].unique()
    uniqueSp2ds = np.union1d(sp2d1, sp2d2)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    sumAllUang = 0
    countUniqueSp2d = len(uniqueSp2ds)
    
    boldFont = Font(bold=True)
    headerFill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    sectionFill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thinBorder = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    def applyHeaderStyle(ws, rowIdx):
        for cell in ws[rowIdx]:
            if cell.value is not None:
                cell.font = boldFont
                cell.fill = headerFill
                cell.border = thinBorder
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def applyDataStyle(ws, rowIdx, isNumberCol=None):
        for idx, cell in enumerate(ws[rowIdx]):
            if cell.value is not None:
                cell.border = thinBorder
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if isNumberCol and (idx + 1) == isNumberCol:
                    cell.number_format = '#,##0'
                    
    for sp2d in uniqueSp2ds:
        ws = wb.create_sheet(title=sp2d)
        d1 = df1[df1['NO SP2D_clean'] == sp2d]
        d2 = df2[df2['Nomor SP2D_clean'] == sp2d]
        
        sumUang = d1['JUMLAH'].sum()
        sumAllUang += sumUang
        
        ws.append(["DATA RINGKASAN"])
        ws[ws.max_row][0].font = boldFont
        ws[ws.max_row][0].fill = sectionFill
        
        ws.append(["NAMA PENERIMA", "NO SP2D", "JUMLAH"])
        applyHeaderStyle(ws, ws.max_row)
        
        if d1.empty:
            ws.append(["(Tidak ada data di sheet Ringkasan)", "", ""])
            applyDataStyle(ws, ws.max_row)
        else:
            for _, row in d1.iterrows():
                ws.append([row['NAMA PENERIMA'], str(row['NO SP2D']), row['JUMLAH']])
                applyDataStyle(ws, ws.max_row, isNumberCol=3)
            
        ws.append([])
        
        ws.append(["DATA SURAT RETUR"])
        ws[ws.max_row][0].font = boldFont
        ws[ws.max_row][0].fill = sectionFill
        
        ws.append(["Nomor SP2D", "Uraian SP2D Retur", "Nilai SP2D"])
        applyHeaderStyle(ws, ws.max_row)
        
        if d2.empty:
            ws.append([sp2d, "", ""])
            applyDataStyle(ws, ws.max_row)
        else:
            for _, row in d2.iterrows():
                uraianVal = row.get('Perbaikan Data Penerima Pembayaran', '')
                if pd.isna(uraianVal) or str(uraianVal).strip() == '':
                    uraianVal = row.get('Uraian SP2D Retur', '')
                ws.append([str(row['Nomor SP2D']), uraianVal, row.get('Nilai SP2D', '')])
                applyDataStyle(ws, ws.max_row)
            
        ws.append([])
        
        ws.append(["DATA TOTAL SP2D"])
        ws[ws.max_row][0].font = boldFont
        ws[ws.max_row][0].fill = sectionFill
        
        ws.append(["NO SP2D", "TOTAL UANG"])
        applyHeaderStyle(ws, ws.max_row)
        
        origSp2d = str(d1['NO SP2D'].iloc[0]) if not d1.empty else sp2d
        ws.append([origSp2d, sumUang])
        applyDataStyle(ws, ws.max_row, isNumberCol=2)
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 35

    for sheetName, totalLabel, totalVal in [
        ("GRAND TOTAL KESELURUHAN", "GRAND TOTAL KESELURUHAN", sumAllUang),
        ("TOTAL SP2D UNIK", "TOTAL SP2D UNIK", countUniqueSp2d)
    ]:
        ws = wb.create_sheet(title=sheetName)
        
        ws.append(["DATA RINGKASAN"])
        ws[ws.max_row][0].font = boldFont
        ws[ws.max_row][0].fill = sectionFill
        ws.append(["NAMA PENERIMA", "NO SP2D", "JUMLAH"])
        applyHeaderStyle(ws, ws.max_row)
        ws.append(["(Tidak ada data di sheet Ringkasan)", "", ""])
        applyDataStyle(ws, ws.max_row)
        ws.append([])
        
        ws.append(["DATA SURAT RETUR"])
        ws[ws.max_row][0].font = boldFont
        ws[ws.max_row][0].fill = sectionFill
        ws.append(["Nomor SP2D", "Uraian SP2D Retur", "Nilai SP2D"])
        applyHeaderStyle(ws, ws.max_row)
        ws.append(["(Tidak ada data di Surat Retur)", "", ""])
        applyDataStyle(ws, ws.max_row)
        ws.append([])
        
        ws.append(["DATA TOTAL SP2D"])
        ws[ws.max_row][0].font = boldFont
        ws[ws.max_row][0].fill = sectionFill
        
        ws.append(["NO SP2D", "TOTAL UANG / JUMLAH SP2D", ""])
        applyHeaderStyle(ws, ws.max_row)
        
        ws.append([totalLabel, totalVal, ""])
        applyDataStyle(ws, ws.max_row, isNumberCol=2)
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 35
        
    wb.save(fileOutput)
    print(f"File berhasil disimpan: {fileOutput}")

if __name__ == "__main__":
    main()
